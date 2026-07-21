"""
Aged Receivables Updater — Vercel Python serverless function.

Core logic: merge a "Source" Aged Receivables export (e.g. Xero-style,
label-spaced columns) into an existing "Last Updated" working sheet
(tight B-G columns, running Total row(s)).

Rules:
1. Match contacts by name (trimmed, case-insensitive) between the two files.
2. For matched contacts: overwrite columns B-F in the target with the
   source's Current / <1 Month / 1 Month / 2 Months / Older values.
   Column G (Total) is left as the existing formula (=B+C+D+E+F) unless
   the row is zeroed (see rule 3).
3. If the computed Total (sum of B:F) for a row is negative, set B-G all
   to 0 for that row.
4. Contacts present in the Source but not in the Target are appended as
   NEW rows at the bottom of the data block (before the Total rows),
   using the same formatting/formula style as existing rows.
5. Contacts present in the Target but not in the Source are left
   completely untouched (no data to update them with).
6. No other rows, columns, sheets, or formatting are touched.
7. The two "Total" summary rows at the bottom are re-pointed to the new
   data range after insertion.
"""

from __future__ import annotations
from copy import copy
from dataclasses import dataclass, field
from io import BytesIO

from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet


class MergeError(Exception):
    """Raised when either file doesn't match the expected Aged Receivables layout."""


@dataclass
class MergeResult:
    output_bytes: bytes
    matched: int = 0
    new_entries: int = 0
    zeroed: int = 0
    unmatched_in_target: int = 0
    sheet_name: str = ""
    warnings: list[str] = field(default_factory=list)


SOURCE_LABELS = {
    "current": "current",
    "<1month": "< 1 month",
    "1month": "1 month",
    "2months": "2 months",
    "older": "older",
    "total": "total",
}

TARGET_LABELS = {
    "contact": "contact",
    "current": "current",
    "<1month": "< 1 month",
    "1month": "1 month",
    "2months": "2 months",
    "older": "older",
    "total": "total",
}


def _norm(v) -> str:
    return str(v).strip().lower() if v is not None else ""


def _find_sheet(wb, preferred_name="Aged Receivables Summary"):
    if preferred_name in wb.sheetnames:
        return wb[preferred_name]
    # fall back: first sheet containing "Contact" somewhere in first 10 rows
    for ws in wb.worksheets:
        for r in range(1, min(15, ws.max_row) + 1):
            for c in range(1, min(15, ws.max_column) + 1):
                if _norm(ws.cell(row=r, column=c).value) == "contact":
                    return ws
    raise MergeError("Could not find an 'Aged Receivables Summary' sheet in the file.")


def _locate_source_columns(ws: Worksheet):
    """Source file: header row has 'Contact' in col A, and labelled columns
    spaced out with blank separator columns (Current, <1 Month, 1 Month,
    2 Months, Older, Total)."""
    header_row = None
    for r in range(1, min(20, ws.max_row) + 1):
        if _norm(ws.cell(row=r, column=1).value) == "contact":
            header_row = r
            break
    if header_row is None:
        raise MergeError("Source file: couldn't find the 'Contact' header row.")

    col_map = {}
    for c in range(1, ws.max_column + 1):
        label = _norm(ws.cell(row=header_row, column=c).value)
        for key, target_label in SOURCE_LABELS.items():
            if label == target_label:
                col_map[key] = c

    required = ["current", "<1month", "1month", "2months", "older", "total"]
    missing = [k for k in required if k not in col_map]
    if missing:
        raise MergeError(f"Source file: missing expected column(s): {missing}")

    # data starts the row after header, ends the row before the "Total" row
    data_start = header_row + 1
    data_end = None
    for r in range(data_start, ws.max_row + 1):
        if _norm(ws.cell(row=r, column=1).value) == "total":
            data_end = r - 1
            break
    if data_end is None:
        raise MergeError("Source file: couldn't find the summary 'Total' row.")

    return header_row, data_start, data_end, col_map


def _locate_target_columns(ws: Worksheet):
    """Target working file: header row has 'Contact' in col A and tight
    columns B..G for Current / <1 Month / 1 Month / 2 Months / Older / Total."""
    header_row = None
    for r in range(1, min(20, ws.max_row) + 1):
        if _norm(ws.cell(row=r, column=1).value) == "contact":
            header_row = r
            break
    if header_row is None:
        raise MergeError("Target file: couldn't find the 'Contact' header row.")

    col_map = {}
    for c in range(1, ws.max_column + 1):
        label = _norm(ws.cell(row=header_row, column=c).value)
        for key, target_label in TARGET_LABELS.items():
            if key != "contact" and label == target_label and key not in col_map:
                col_map[key] = c

    required = ["current", "<1month", "1month", "2months", "older", "total"]
    missing = [k for k in required if k not in col_map]
    if missing:
        raise MergeError(f"Target file: missing expected column(s): {missing}")

    data_start = header_row + 1
    # data ends at the row before the first "Total" row found below the header
    data_end = None
    for r in range(data_start, ws.max_row + 1):
        if _norm(ws.cell(row=r, column=1).value) == "total":
            data_end = r - 1
            break
    if data_end is None:
        raise MergeError("Target file: couldn't find the summary 'Total' row.")

    return header_row, data_start, data_end, col_map


def merge_workbooks(source_file, target_file) -> MergeResult:
    """
    source_file / target_file: file-like objects (bytes or BytesIO) of .xlsx files.
    Returns a MergeResult with the merged workbook bytes and a summary.
    """
    warnings: list[str] = []

    # ---- read source (values only) ----
    src_wb = load_workbook(source_file, data_only=True)
    src_ws = _find_sheet(src_wb)
    _, s_start, s_end, s_cols = _locate_source_columns(src_ws)

    src_data = {}
    src_order = []
    for r in range(s_start, s_end + 1):
        name = src_ws.cell(row=r, column=1).value
        if name is None or str(name).strip() == "":
            continue
        name_key = str(name).strip()
        cur = src_ws.cell(row=r, column=s_cols["current"]).value or 0
        lt1 = src_ws.cell(row=r, column=s_cols["<1month"]).value or 0
        m1 = src_ws.cell(row=r, column=s_cols["1month"]).value or 0
        m2 = src_ws.cell(row=r, column=s_cols["2months"]).value or 0
        older = src_ws.cell(row=r, column=s_cols["older"]).value or 0
        if name_key in src_data:
            warnings.append(f"Duplicate contact in source file: '{name_key}' — last occurrence used.")
        src_data[name_key] = (cur, lt1, m1, m2, older)
        if name_key not in src_order:
            src_order.append(name_key)

    # ---- load target (formulas preserved) ----
    tgt_wb = load_workbook(target_file, data_only=False)
    tgt_ws = _find_sheet(tgt_wb)
    header_row, t_start, t_end, t_cols = _locate_target_columns(tgt_ws)

    B, C, D, E, F, G = (
        t_cols["current"], t_cols["<1month"], t_cols["1month"],
        t_cols["2months"], t_cols["older"], t_cols["total"],
    )

    tgt_name_to_row = {}
    for r in range(t_start, t_end + 1):
        nm = tgt_ws.cell(row=r, column=1).value
        if nm:
            tgt_name_to_row[str(nm).strip()] = r

    matched = 0
    zeroed = 0
    for name, vals in src_data.items():
        r = tgt_name_to_row.get(name)
        if r is None:
            continue
        cur, lt1, m1, m2, older = vals
        total = cur + lt1 + m1 + m2 + older
        if total < 0:
            for col in (B, C, D, E, F):
                tgt_ws.cell(row=r, column=col).value = 0
            zeroed += 1
        else:
            tgt_ws.cell(row=r, column=B).value = cur
            tgt_ws.cell(row=r, column=C).value = lt1
            tgt_ws.cell(row=r, column=D).value = m1
            tgt_ws.cell(row=r, column=E).value = m2
            tgt_ws.cell(row=r, column=F).value = older
        matched += 1

    new_entries = [n for n in src_order if n not in tgt_name_to_row]
    unmatched_in_target = len([n for n in tgt_name_to_row if n not in src_data])

    n_new = len(new_entries)
    insert_at = t_end + 1
    if n_new:
        tgt_ws.insert_rows(insert_at, amount=n_new)

    template_row = t_start
    style_cells = {col: tgt_ws.cell(row=template_row, column=col) for col in range(1, G + 1)}

    for i, name in enumerate(new_entries):
        r = insert_at + i
        cur, lt1, m1, m2, older = src_data[name]
        total = cur + lt1 + m1 + m2 + older
        if total < 0:
            cur = lt1 = m1 = m2 = older = 0
            zeroed += 1

        tgt_ws.cell(row=r, column=1).value = name
        tgt_ws.cell(row=r, column=B).value = cur
        tgt_ws.cell(row=r, column=C).value = lt1
        tgt_ws.cell(row=r, column=D).value = m1
        tgt_ws.cell(row=r, column=E).value = m2
        tgt_ws.cell(row=r, column=F).value = older
        g_letter = get_column_letter(G)
        b_letter, c_letter, d_letter, e_letter, f_letter = (
            get_column_letter(B), get_column_letter(C), get_column_letter(D),
            get_column_letter(E), get_column_letter(F),
        )
        tgt_ws.cell(row=r, column=G).value = (
            f"={b_letter}{r}+{c_letter}{r}+{d_letter}{r}+{e_letter}{r}+{f_letter}{r}"
        )

        for col in range(1, G + 1):
            src_style_cell = style_cells[col]
            dst = tgt_ws.cell(row=r, column=col)
            dst.font = copy(src_style_cell.font)
            dst.number_format = src_style_cell.number_format
            dst.alignment = copy(src_style_cell.alignment)
            dst.border = copy(src_style_cell.border)
            dst.fill = copy(src_style_cell.fill)
        tgt_ws.row_dimensions[r].height = tgt_ws.row_dimensions[template_row].height

    new_last_data_row = t_end + n_new

    # ---- fix the Total row(s) below the data block ----
    total_rows = []
    for r in range(insert_at, tgt_ws.max_row + 1):
        if _norm(tgt_ws.cell(row=r, column=1).value) == "total":
            total_rows.append(r)

    if total_rows:
        first_total_row = total_rows[0]
        for col in (B, C, D, E, F, G):
            col_letter = get_column_letter(col)
            cell = tgt_ws.cell(row=first_total_row, column=col)
            existing = cell.value
            if isinstance(existing, str) and existing.upper().startswith("=SUM("):
                cell.value = f"=SUM({col_letter}{t_start}:{col_letter}{new_last_data_row})"
            else:
                cell.value = f"=SUM({col_letter}{t_start}:{col_letter}{new_last_data_row})"

        for extra_row in total_rows[1:]:
            for col in (B, C, D, E, F, G):
                col_letter = get_column_letter(col)
                tgt_ws.cell(row=extra_row, column=col).value = f"={col_letter}{first_total_row}"
    else:
        warnings.append("No 'Total' row found below the data block to update — totals were left as-is.")

    buf = BytesIO()
    tgt_wb.save(buf)
    buf.seek(0)

    return MergeResult(
        output_bytes=buf.read(),
        matched=matched,
        new_entries=n_new,
        zeroed=zeroed,
        unmatched_in_target=unmatched_in_target,
        sheet_name=tgt_ws.title,
        warnings=warnings,
    )


# ---------------------------------------------------------------------------
# ASGI endpoint — Vercel serves this file at /api/merge; the page at
# /ar-update/ posts the two files here. Routes are registered under both the
# full path and the root so the app works regardless of path forwarding.
# ---------------------------------------------------------------------------
from fastapi import FastAPI, UploadFile, File, HTTPException
from fastapi.responses import StreamingResponse, JSONResponse

app = FastAPI(title="Aged Receivables Updater")


async def _do_merge(source: UploadFile, target: UploadFile):
    for f in (source, target):
        if not f.filename.lower().endswith((".xlsx", ".xlsm")):
            raise HTTPException(400, f"'{f.filename}' is not an .xlsx/.xlsm file.")

    source_bytes = await source.read()
    target_bytes = await target.read()

    try:
        result = merge_workbooks(BytesIO(source_bytes), BytesIO(target_bytes))
    except MergeError as e:
        raise HTTPException(422, str(e))
    except Exception as e:
        raise HTTPException(500, f"Unexpected error while processing the files: {e}")

    out_name = target.filename.rsplit(".", 1)[0] + "_updated.xlsx"

    headers = {
        "Content-Disposition": f'attachment; filename="{out_name}"',
        "X-Matched": str(result.matched),
        "X-New-Entries": str(result.new_entries),
        "X-Zeroed": str(result.zeroed),
        "X-Unmatched-In-Target": str(result.unmatched_in_target),
        "X-Warnings": " | ".join(result.warnings) if result.warnings else "",
        "Access-Control-Expose-Headers": "X-Matched, X-New-Entries, X-Zeroed, X-Unmatched-In-Target, X-Warnings",
    }

    return StreamingResponse(
        BytesIO(result.output_bytes),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers=headers,
    )


@app.post("/api/merge")
async def merge_full_path(source: UploadFile = File(...), target: UploadFile = File(...)):
    return await _do_merge(source, target)


@app.post("/")
async def merge_root(source: UploadFile = File(...), target: UploadFile = File(...)):
    return await _do_merge(source, target)


@app.get("/api/merge")
@app.get("/")
def health():
    return JSONResponse({"status": "ok", "tool": "aged-receivables-updater"})
